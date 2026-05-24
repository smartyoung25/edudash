"""
주간보고 양식의 (교육생별) 운영현황 시트를 채움.
출석시간(분)은 Drive '전체_팀_운영_현황' xlsx의 회차별 교육시간 범위에서 계산.
학생 출결은 같은 파일의 R5 '불참자' 명단(코디가 직접 관리하는 ground truth).
"""
import sys, io, sqlite3, re, datetime, shutil, warnings
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')
warnings.filterwarnings('ignore')
from openpyxl import load_workbook

SRC_FORM = "C:/Users/IIamHub2/Documents/네이트온 받은 파일/2026년 맞춤형과정 주간보고_이암허브.xlsx"
DRIVE = "data/drive-team-status.xlsx"
today = datetime.date.today().strftime("%Y%m%d")
OUT = f"data/2026년_맞춤형과정_주간보고_이암허브_{today}.xlsx"
shutil.copy(SRC_FORM, OUT)

# 1) Drive 파일에서 팀별 회차 정보 추출
def parse_minutes(s):
    """'09:00~17:00' 또는 '08:00 - 17:00' → 분"""
    if not s: return None
    m = re.search(r'(\d{1,2}):(\d{2})\s*[~\-–]\s*(\d{1,2}):(\d{2})', str(s))
    if not m: return None
    h1, m1, h2, m2 = map(int, m.groups())
    total = (h2*60+m2) - (h1*60+m1)
    # 점심시간 60분 차감 (실교육시간)
    return max(0, total - 60) if total > 60 else total

def parse_date(v, year=2026):
    if v is None or v == '': return None
    if isinstance(v, datetime.datetime): return v.date()
    if isinstance(v, datetime.date): return v
    s = str(v).strip()
    m = re.match(r'(\d{1,2})[월\.\-/](\d{1,2})', s)
    if m: return datetime.date(year, int(m.group(1)), int(m.group(2)))
    m = re.match(r'(\d{4})-(\d{1,2})-(\d{1,2})', s)
    if m: return datetime.date(int(m.group(1)), int(m.group(2)), int(m.group(3)))
    return None

def parse_absentees(s):
    if not s or str(s).strip().upper() == 'X': return set()
    text = str(s)
    # 괄호 안 사유 제거: '(오대철-병원진료, 김경훈-자녀질병)' → '오대철, 김경훈'
    text = re.sub(r'\([^)]*\)', lambda m: re.sub(r'-[^,)]+', '', m.group()), text)
    text = text.replace('(','').replace(')','')
    names = re.split(r'[,\s/]+', text)
    return {n.strip() for n in names if n.strip() and n.strip().upper() != 'X' and re.match(r'^[가-힣]{2,4}$', n.strip())}

drive_wb = load_workbook(DRIVE, data_only=True)
team_sheets = {}
for sname in drive_wb.sheetnames:
    ws = drive_wb[sname]
    sessions = {}  # session_no -> {minutes, date, absentees}
    for col in range(2, ws.max_column+1):
        no = ws.cell(1, col).value
        m = re.search(r'(\d+)', str(no or ''))
        if not m: continue
        session_no = int(m.group(1))
        mins = parse_minutes(ws.cell(2, col).value)
        d = parse_date(ws.cell(3, col).value)
        absent = parse_absentees(ws.cell(5, col).value)
        if mins is None and d is None and not absent: continue
        sessions[session_no] = {'minutes': mins, 'date': d, 'absentees': absent}
    team_sheets[sname] = sessions

print("Drive 팀 시트:", len(team_sheets))
for name, ss in team_sheets.items():
    valid = [k for k,v in ss.items() if v['minutes'] or v['date']]
    print(f"  {name}: 회차정보 {len(valid)}개")

# 2) DB 팀·회원 로드
con = sqlite3.connect("data/app.db")
con.row_factory = sqlite3.Row
teams = list(con.execute("SELECT id, name, product, cohort FROM teams"))
members = list(con.execute("SELECT id, team_id, name, edu_status FROM members"))

def team_course_keys(t):
    m = re.search(r'\d+', str(t['cohort']))
    if not m: return []
    n = int(m.group())
    return [f"{t['product']}{n:02d}기", f"{t['product']} {n:02d}기", f"{t['product']}{n}기", f"{t['product']} {n}기"]

# 양식 과정번호 → DB 팀 매핑
team_by_course = {}
for t in teams:
    for k in team_course_keys(t):
        team_by_course[k] = t

# DB 팀명 → drive sheet 매핑 (팀명이 일치하거나 부분일치)
def find_drive_sheet(team_name):
    if team_name in team_sheets: return team_name
    for sn in team_sheets:
        if team_name in sn or sn in team_name: return sn
    return None

# 3) 양식 채움
wb = load_workbook(OUT)
ws = wb['(교육생별) 운영현황']
COL_COURSE, COL_NAME = 1, 3
COL_PROGRESS, COL_ATTRATE = 45, 46  # AS, AT

updated_students = 0; unmatched_students = 0; total_attendance_cells = 0
for row in range(3, ws.max_row+1):
    course_no = ws.cell(row, COL_COURSE).value
    name = ws.cell(row, COL_NAME).value
    if not course_no or not name: continue
    team = team_by_course.get(str(course_no).strip())
    if not team:
        unmatched_students += 1; continue
    sheet_name = find_drive_sheet(team['name'])
    if not sheet_name:
        unmatched_students += 1; continue
    sessions = team_sheets[sheet_name]
    sname = str(name).strip()

    total_planned = 0; attended_min = 0
    for session_no, info in sessions.items():
        if session_no < 1 or session_no > 40: continue
        date_col = 47 + (session_no-1)*2
        time_col = date_col + 1
        if info['date']:
            ws.cell(row, date_col).value = info['date']
        if info['minutes'] is not None:
            absent = sname in info['absentees']
            mins = 0 if absent else info['minutes']
            ws.cell(row, time_col).value = mins
            total_planned += info['minutes']
            attended_min += mins
            total_attendance_cells += 1

    # 출석률 = 실제 출석 분 / 예정 분
    if total_planned > 0:
        ws.cell(row, COL_ATTRATE).value = round(attended_min/total_planned*100)
    # 진행률 = 회차정보 있는 회차 / 양식상 교육일수 (R = 17번째 컬럼?)
    total_days = ws.cell(row, 17).value  # Q 교육일수
    done = sum(1 for v in sessions.values() if v['minutes'] is not None)
    if total_days and isinstance(total_days, (int, float)) and total_days > 0:
        ws.cell(row, COL_PROGRESS).value = round(done/total_days*100)
    else:
        ws.cell(row, COL_PROGRESS).value = done  # 회차 수만 기록
    updated_students += 1

wb.save(OUT)
print(f"\n✓ 출력: {OUT}")
print(f"  학생 채움 {updated_students}건, 미매칭 {unmatched_students}건, 출석셀 {total_attendance_cells}개 갱신")
