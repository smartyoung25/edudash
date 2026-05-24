import sys, io, warnings
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')
warnings.filterwarnings('ignore')
from openpyxl import load_workbook
wb = load_workbook("data/drive-team-status.xlsx", data_only=True)
for name in ["딸기15 논산", "감귤4", "한우7기"]:
    ws = wb[name]
    print(f"\n========== {name} max_row={ws.max_row} max_col={ws.max_column} ==========")
    rows = list(ws.iter_rows(values_only=True))
    # 비어있지 않은 행만 표시
    for i, row in enumerate(rows, 1):
        if not any(v not in (None, '', 0) for v in row): continue
        vals = list(row)[:16]
        print(f" R{i}", [str(v)[:18] if v is not None else '' for v in vals])
        if i > 60: break
