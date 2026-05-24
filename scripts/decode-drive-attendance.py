"""tool-results의 base64 JSON을 디코딩해 xlsx로 저장 + 구조 분석"""
import sys, io, json, base64, warnings
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')
warnings.filterwarnings('ignore')

src = r"C:\Users\IIamHub2\.claude\projects\C--2026-edu-preject\38d73d04-9144-4ea3-88dd-ec6b71a9967d\tool-results\mcp-dc84a135-13a1-4610-8646-c0f5c9485e7f-download_file_content-1779598553196.txt"
with open(src, 'r', encoding='utf-8') as f:
    data = json.load(f)
print("title:", data.get('title'), "mime:", data.get('mimeType'), "len:", len(data.get('content','')))
out_path = "data/drive-team-status.xlsx"
with open(out_path, 'wb') as f:
    f.write(base64.b64decode(data['content']))
print("saved:", out_path)

from openpyxl import load_workbook
wb = load_workbook(out_path, data_only=True, read_only=True)
print("SHEETS:", wb.sheetnames)
for name in wb.sheetnames[:5]:
    ws = wb[name]
    print(f"\n=== {name} dim={ws.dimensions} max_row={ws.max_row} max_col={ws.max_column}")
    for i, row in enumerate(ws.iter_rows(max_row=6, values_only=True), 1):
        vals = list(row)[:20]
        print(f" R{i}", [str(v)[:20] if v is not None else '' for v in vals])
