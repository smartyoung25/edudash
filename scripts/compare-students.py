"""사이트 DB(members) vs 주간보고 양식의 교육생 인원 비교"""
import sys, io, sqlite3, re
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')

import warnings
warnings.filterwarnings('ignore')
from openpyxl import load_workbook

TEMPLATE = "C:/Users/IIamHub2/Documents/네이트온 받은 파일/2026년 맞춤형과정 주간보고_이암허브.xlsx"
DB = "data/app.db"

# --- 양식에서 학생 추출 (과정번호, 성명, 현재상태)
wb = load_workbook(TEMPLATE, data_only=True, read_only=True)
ws = wb['(교육생별) 운영현황']
tmpl_by_course = {}  # course_no -> [(name, status), ...]
for i, row in enumerate(ws.iter_rows(min_row=3, values_only=True), 3):
    course_no = row[0]; name = row[2]; status = row[22] if len(row) > 22 else None
    if not course_no or not name: continue
    course_no = str(course_no).strip()
    name = str(name).strip()
    tmpl_by_course.setdefault(course_no, []).append((name, str(status or '').strip()))
wb.close()

# --- DB에서 학생 추출 (팀명 + 기수 + 학생목록)
con = sqlite3.connect(DB)
con.row_factory = sqlite3.Row
teams = list(con.execute("SELECT id, name, product, cohort FROM teams ORDER BY id"))
members = list(con.execute("SELECT team_id, name, edu_status FROM members ORDER BY team_id, id"))

db_by_team = {}
for m in members:
    db_by_team.setdefault(m['team_id'], []).append((m['name'], m['edu_status'] or ''))

# 팀 → 양식 과정번호 매칭 (e.g. "감귤7기" → "감귤 07기" / "감귤07기")
def team_to_course_keys(t):
    product = t['product']
    cohort = str(t['cohort']).strip()
    # cohort에서 숫자만
    m = re.search(r'\d+', cohort)
    num = int(m.group()) if m else None
    if num is None: return []
    keys = []
    keys.append(f"{product}{num:02d}기")
    keys.append(f"{product} {num:02d}기")
    keys.append(f"{product}{num}기")
    keys.append(f"{product} {num}기")
    return keys

print("=" * 100)
print(f"{'팀명':30s} | {'기수':6s} | {'사이트':6s} | {'양식':6s} | {'양식과정번호':16s} | 일치여부")
print("-" * 100)

total_site = 0; total_form = 0
issues = []
for t in teams:
    keys = team_to_course_keys(t)
    matched_key = None
    form_list = []
    for k in keys:
        if k in tmpl_by_course:
            matched_key = k; form_list = tmpl_by_course[k]; break
    site_list = db_by_team.get(t['id'], [])
    site_active = [n for n, s in site_list if s != '교육취소']
    n_site = len(site_active)
    n_form = len(form_list)
    total_site += n_site; total_form += n_form
    mark = '✅' if n_site == n_form else '⚠️ 차이'
    print(f"{t['name'][:30]:30s} | {t['cohort'][:6]:6s} | {n_site:6d} | {n_form:6d} | {(matched_key or '미매칭'):16s} | {mark}")
    if n_site != n_form:
        s_set = set(n for n, _ in site_active and [(x,'') for x in site_active] or [])
        site_names = set(site_active)
        form_names = set(n for n, _ in form_list)
        only_site = site_names - form_names
        only_form = form_names - site_names
        if only_site: issues.append(f"  [{t['name']}] 사이트에만: {sorted(only_site)}")
        if only_form: issues.append(f"  [{t['name']}] 양식에만:   {sorted(only_form)}")

print("-" * 100)
print(f"{'합계':30s} | {'':6s} | {total_site:6d} | {total_form:6d}")

if issues:
    print("\n=== 차이 상세 ===")
    for line in issues: print(line)

# 양식에 있는데 우리 시스템에 팀이 없는 과정
db_keys = set()
for t in teams:
    for k in team_to_course_keys(t):
        if k in tmpl_by_course: db_keys.add(k)
extra = sorted(set(tmpl_by_course.keys()) - db_keys)
if extra:
    print(f"\n=== 사이트에 팀이 없는 양식 과정 ({len(extra)}개) ===")
    for k in extra:
        print(f"  {k}: {len(tmpl_by_course[k])}명")
