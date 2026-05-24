"""
주간보고 양식의 (교육생별) 운영현황 시트에서
매칭되는 교육생 행에 진행률·출석률·회차 출석여부를 채워서 신규 파일로 저장.
원본 양식의 컬럼/포맷/메타데이터는 그대로 유지.
"""
import sys, io, sqlite3, re, datetime, shutil
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')

import warnings
warnings.filterwarnings('ignore')
from openpyxl import load_workbook

SRC = "C:/Users/IIamHub2/Documents/네이트온 받은 파일/2026년 맞춤형과정 주간보고_이암허브.xlsx"
DB = "data/app.db"
today = datetime.date.today().strftime("%Y%m%d")
OUT = f"data/2026년_맞춤형과정_주간보고_이암허브_{today}.xlsx"

shutil.copy(SRC, OUT)

con = sqlite3.connect(DB)
con.row_factory = sqlite3.Row
teams = list(con.execute("SELECT id, name, product, cohort, total_sessions FROM teams"))
members = list(con.execute("SELECT id, team_id, name, edu_status FROM members"))
sessions = list(con.execute("SELECT id, team_id, session_no, scheduled_date, status FROM sessions"))
attendance = list(con.execute("SELECT session_id, member_id, status FROM attendance"))

# 인덱스 정리
sess_by_id = {s['id']: s for s in sessions}
sess_by_team_no = {}
for s in sessions:
    sess_by_team_no.setdefault(s['team_id'], {})[s['session_no']] = s
att_by_member = {}
for a in attendance:
    att_by_member.setdefault(a['member_id'], []).append(a)

def team_course_keys(t):
    m = re.search(r'\d+', str(t['cohort']))
    if not m: return []
    n = int(m.group())
    return [f"{t['product']}{n:02d}기", f"{t['product']} {n:02d}기", f"{t['product']}{n}기", f"{t['product']} {n}기"]

team_by_key = {}
for t in teams:
    for k in team_course_keys(t):
        team_by_key.setdefault(k, t)

# 진행률·출석률 산출
def team_progress(team_id):
    ss = sess_by_team_no.get(team_id, {})
    if not ss: return 0
    done = sum(1 for s in ss.values() if s['status'] == 'done')
    return round(done / len(ss) * 100) if ss else 0

def member_attendance_rate(member_id, team_id):
    ss = sess_by_team_no.get(team_id, {})
    valid_ids = {s['id'] for s in ss.values()}
    rows = [a for a in att_by_member.get(member_id, []) if a['session_id'] in valid_ids]
    if not rows: return None
    pres = sum(1 for a in rows if a['status'] == 'present')
    return round(pres / len(rows) * 100)

def member_session_attendance(member_id, team_id, session_no):
    """해당 회차에 출석했으면 '출석', 결석이면 '결석', 미기록이면 None"""
    ss = sess_by_team_no.get(team_id, {}).get(session_no)
    if not ss: return None
    for a in att_by_member.get(member_id, []):
        if a['session_id'] == ss['id']:
            return '출석' if a['status'] == 'present' else '결석'
    return None

# 양식 열기 (data_only=False 로 수식 보존)
wb = load_workbook(OUT)
ws = wb['(교육생별) 운영현황']

# 컬럼 좌표
COL_COURSE = 1   # A 과정번호
COL_NAME = 3     # C 성명
COL_PROGRESS = 45  # AS 진행률
COL_ATTRATE = 46   # AT 출석률
# 1일차 교육일자 AU=47, 출석시간 AV=48; 일차마다 2칸씩 증가, 최대 40일차

updated = 0; unmatched = 0
for row in range(3, ws.max_row + 1):
    course_no = ws.cell(row, COL_COURSE).value
    name = ws.cell(row, COL_NAME).value
    if not course_no or not name: continue
    key = str(course_no).strip()
    team = team_by_key.get(key)
    if not team:
        unmatched += 1; continue
    mname = str(name).strip()
    member = next((m for m in members if m['team_id'] == team['id'] and m['name'] == mname), None)
    if not member:
        unmatched += 1; continue

    # 진행률·출석률
    ws.cell(row, COL_PROGRESS).value = team_progress(team['id'])
    ar = member_attendance_rate(member['id'], team['id'])
    if ar is not None: ws.cell(row, COL_ATTRATE).value = ar

    # 회차별 — 교육일자 + 출석상태(원래는 출석시간이지만 시간 데이터 없음 → '출석'/'결석'만)
    for day_idx in range(1, 41):
        date_col = 47 + (day_idx - 1) * 2
        time_col = date_col + 1
        ss = sess_by_team_no.get(team['id'], {}).get(day_idx)
        if not ss: continue
        try:
            ws.cell(row, date_col).value = datetime.datetime.strptime(ss['scheduled_date'], '%Y-%m-%d').date()
        except Exception:
            ws.cell(row, date_col).value = ss['scheduled_date']
        att = member_session_attendance(member['id'], team['id'], day_idx)
        if att is not None:
            ws.cell(row, time_col).value = att
    updated += 1

wb.save(OUT)
print(f"✓ 출력: {OUT}")
print(f"   채움 {updated}건, 미매칭 {unmatched}건")
